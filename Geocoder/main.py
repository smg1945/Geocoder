import openpyxl
import requests
from urllib.parse import urlparse
import win32com.client


workbook = openpyxl.load_workbook(filename="C:/Users/Admin/Desktop/data.xlsx")
ws = workbook.active
data_list = []
excel = win32com.client.Dispatch('Excel.Application')
excel.Visible = True
winwb = excel.Workbooks.Open('C:/Users/Admin/Desktop/complete.xlsx')
winws = winwb.activesheet

j = 1

for i in range(2, ws.max_row + 1):
    url = 'https://dapi.kakao.com/v2/local/search/address.json?&query=' + ws[f'A{i}'].value
    result = requests.get(urlparse(url).geturl(), headers={'Authorization': "KakaoAK 0cff6ee43a1d4da0588082ffd099317d"}).json()
    try:
        lat = result['documents'][0]['y']
        lon = result['documents'][0]['x']
        winws.cells(j,1).value = ws[f'A{j}'].value
        winws.cells(j,2).value = lat
        winws.cells(j,3).value = lon
        print(f'{winws.cells(j,1).value} 추출 성공')
    except:
        winws.cells(j,1).value = ws[f'A{j}'].value
        winws.cells(j,2).value = ''
        winws.cells(j,3).value = ''
        print(f'{winws.cells(j,1).value} 추출 실패')
    j += 1
winwb.Save()